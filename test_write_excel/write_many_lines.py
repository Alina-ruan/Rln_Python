import openpyxl
'''
写入多行多列数据
'''

#读取文件
'''
def read_data():
    datas = open('./test_data.txt', 'r', encoding='utf-8')
    data = datas.readlines()
    for i in data:
        print(i)

read_data()
'''

#写入多行文件
def write_many_lines():
    work_book = openpyxl.Workbook()
    sheet = work_book.create_sheet('new')
    data = open(r'./test_data.txt', 'r', encoding='utf-8')
    datas = data.readlines()
    #通过enumerate方法来给每一行创建索引
    for index, row in enumerate(datas):
        #print(index, row.split())
        d = row.split()
        for col in range(len(d)):
            sheet.cell(index+1, col+1, d[col])
    work_book.save('这是写入多行的测试文件.xlsx')

if __name__ == '__main__':
    write_many_lines()