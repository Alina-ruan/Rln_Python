import openpyxl
# 写入数据到excel文件

'''
1、创建对象：调用对象的一些属性和方法
2、创建sheet
3、准备数据
4、几行几列：确定如何写入
5、保存文件
'''

def write_excel():
    # 创建表格对象
    excel = openpyxl.Workbook()
    # 创建sheet
    sheet = excel.create_sheet('sheet1')
    # 准备写入的数据
    list1 = ['11', 'aa', 'bb']
    # 循环读取文件的值写入
    for i in range(len(list1)):
        sheet.cell(1, i+1, list1[i])

    # 保存文件
    excel.save('这是第一个测试文件.xlsx')

if __name__ == '__main__':
    write_excel()


