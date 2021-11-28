import openpyxl

def write_append_lines():
    # 读取excel文件
    excel = openpyxl.load_workbook('这是写入多行的测试文件.xlsx')
    # 读取sheet
    sheet = excel['new']
    # 获取文件的最大行，用来确定从哪一行开始追加写入数据
    max_rows = sheet.max_row
    print(max_rows)
    # 准备数据
    list = ['新名字', '新密码']
    # 遍历列表数据
    for d in range(len(list)):
        sheet.cell(max_rows+1, d+1, list[d])

    excel.save('这是写入多行的测试文件.xlsx')

if __name__ == '__main__':
    write_append_lines()

